from dataclasses import dataclass
from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError as DjangoValidationError
from django.core.validators import validate_email
from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from apps.core.choices import GenderChoices, GradeChoices, RoleChoices
from apps.dicts.models import ClassDict
from apps.students.models import StudentProfile


HEADER_ALIASES = {
    'username': {'username', '用户名', '账号'},
    'stu_no': {'stu_no', 'student_no', '学号'},
    'realname': {'realname', 'name', '姓名', '真实姓名'},
    'phone': {'phone', '手机号', '手机', '联系电话'},
    'email': {'email', '邮箱'},
    'address': {'address', '地址', '家庭住址'},
    'age': {'age', '年龄'},
    'gender': {'gender', '性别'},
    'class_id': {'class_id', '班级id', '班级ID'},
    'grade': {'grade', '年级'},
    'class_name': {'class_name', '班级', '班级名称'},
    'password': {'password', '密码', '初始密码'},
}

REQUIRED_COLUMNS = {'stu_no', 'realname'}
DEFAULT_PASSWORD = 'z123456.'
TEMPLATE_HEADERS = [
    '学号',
    '姓名',
    '性别',
    '年龄',
    '手机号',
    '邮箱',
    '年级',
    '班级',
    '班级ID',
    '用户名',
    '密码',
    '地址',
]


@dataclass
class ImportRow:
    row_number: int
    username: str
    stu_no: str
    realname: str
    phone: str = ''
    email: str = ''
    address: str = ''
    age: int | None = None
    gender: str = ''
    class_obj: ClassDict | None = None
    password: str = DEFAULT_PASSWORD


def _normalize_header(value):
    return str(value or '').strip()


def _normalize_cell(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _build_header_map(header_row):
    alias_to_field = {
        alias.lower(): field
        for field, aliases in HEADER_ALIASES.items()
        for alias in aliases
    }
    header_map = {}
    for index, value in enumerate(header_row):
        header = _normalize_header(value)
        if not header:
            continue
        field = alias_to_field.get(header.lower())
        if field:
            header_map[field] = index
    return header_map


def _normalize_gender(value):
    gender = value.strip()
    if not gender:
        return ''

    gender_map = {
        '男': GenderChoices.MALE,
        '女': GenderChoices.FEMALE,
        'male': GenderChoices.MALE,
        'female': GenderChoices.FEMALE,
        'm': GenderChoices.MALE,
        'f': GenderChoices.FEMALE,
        'MALE': GenderChoices.MALE,
        'FEMALE': GenderChoices.FEMALE,
    }
    return gender_map.get(gender, '')


def _parse_age(value, row_number, errors):
    if not value:
        return None
    try:
        age = int(value)
    except ValueError:
        errors.append({'row': row_number, 'field': 'age', 'message': '年龄必须是整数'})
        return None

    if age < 1 or age > 150:
        errors.append({'row': row_number, 'field': 'age', 'message': '年龄必须在 1 到 150 之间'})
        return None
    return age


def _resolve_class(row_data, row_number, classes_by_id, classes_by_grade_name, errors):
    class_id = row_data.get('class_id', '')
    grade = row_data.get('grade', '')
    class_name = row_data.get('class_name', '')

    if class_id:
        try:
            parsed_class_id = int(class_id)
        except ValueError:
            errors.append({'row': row_number, 'field': 'class_id', 'message': '班级ID必须是整数'})
            return None

        class_obj = classes_by_id.get(parsed_class_id)
        if class_obj is None:
            errors.append({'row': row_number, 'field': 'class_id', 'message': '班级ID不存在'})
        return class_obj

    if not grade and not class_name:
        return None

    if grade not in GradeChoices.values:
        errors.append({'row': row_number, 'field': 'grade', 'message': '年级编码无效'})
        return None

    if not class_name:
        errors.append({'row': row_number, 'field': 'class_name', 'message': '按年级匹配班级时必须填写班级名称'})
        return None

    class_obj = classes_by_grade_name.get((grade, class_name))
    if class_obj is None:
        errors.append({'row': row_number, 'field': 'class_name', 'message': '年级和班级名称未匹配到班级'})
    return class_obj


def _row_to_dict(row, header_map):
    return {
        field: _normalize_cell(row[index] if index < len(row) else '')
        for field, index in header_map.items()
    }


def _format_errors(errors, row_context=None):
    row_context = row_context or {}
    grouped_errors = {}

    for error in errors:
        row_number = error['row']
        context = row_context.get(row_number, {})
        row_error = grouped_errors.setdefault(row_number, {
            'row': row_number,
            'stu_no': context.get('stu_no', ''),
            'realname': context.get('realname', ''),
            'messages': [],
        })
        message = error['message']
        if message not in row_error['messages']:
            row_error['messages'].append(message)

    return list(grouped_errors.values())


def build_student_import_template():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '学生导入模板'
    worksheet.append(TEMPLATE_HEADERS)
    worksheet.append([
        '20260001',
        '张三',
        '男',
        13,
        '13800000001',
        'zhangsan@example.com',
        'GRADE_7',
        '1班',
        '',
        '',
        '',
        '海口市示例路 1 号',
    ])
    worksheet.append([
        '20260002',
        '李四',
        '女',
        13,
        '13800000002',
        'lisi@example.com',
        'GRADE_7',
        '1班',
        '',
        '',
        '',
        '海口市示例路 2 号',
    ])

    header_fill = PatternFill(fill_type='solid', fgColor='D9EAF7')
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    column_widths = [14, 12, 8, 8, 16, 24, 14, 12, 10, 16, 16, 28]
    for index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = width

    worksheet.freeze_panes = 'A2'

    help_sheet = workbook.create_sheet('填写说明')
    help_rows = [
        ['项目', '说明'],
        ['必填列', '学号、姓名'],
        ['用户名', '不填时默认使用学号作为用户名'],
        ['密码', f'不填时使用接口参数 default_password，默认 {DEFAULT_PASSWORD}'],
        ['性别', '支持：男、女、MALE、FEMALE'],
        ['班级匹配', '优先使用班级ID；未填班级ID时，使用“年级 + 班级”匹配'],
        ['年级编码', ', '.join(GradeChoices.values)],
        ['导入策略', '先校验整张表；任意行有错误则不写入任何数据'],
    ]
    for row in help_rows:
        help_sheet.append(row)
    help_sheet.column_dimensions['A'].width = 16
    help_sheet.column_dimensions['B'].width = 90
    for cell in help_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output.getvalue()


def import_students_from_excel(file_obj, default_password=DEFAULT_PASSWORD):
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return {
            'total_rows': 0,
            'imported_count': 0,
            'failed_count': 0,
            'errors': _format_errors([
                {'row': 1, 'field': 'file', 'message': 'Excel不能为空'},
            ]),
        }

    header_map = _build_header_map(rows[0])
    missing_columns = REQUIRED_COLUMNS - set(header_map)
    if missing_columns:
        errors = [
            {'row': 1, 'field': column, 'message': '缺少必填列'}
            for column in sorted(missing_columns)
        ]
        return {
            'total_rows': 0,
            'imported_count': 0,
            'failed_count': 0,
            'errors': _format_errors(errors),
        }

    User = get_user_model()
    classes = ClassDict.objects.all()
    classes_by_id = {class_obj.id: class_obj for class_obj in classes}
    classes_by_grade_name = {
        (class_obj.grade, class_obj.name): class_obj
        for class_obj in classes
    }
    existing_usernames = set(User.objects.values_list('username', flat=True))
    existing_stu_nos = set(StudentProfile.objects.values_list('stu_no', flat=True))
    seen_usernames = set()
    seen_stu_nos = set()
    import_rows = []
    errors = []
    row_context = {}

    for row_number, row in enumerate(rows[1:], start=2):
        row_data = _row_to_dict(row, header_map)
        if not any(row_data.values()):
            continue

        stu_no = row_data.get('stu_no', '')
        realname = row_data.get('realname', '')
        username = row_data.get('username', '') or stu_no
        row_context[row_number] = {
            'stu_no': stu_no,
            'realname': realname,
        }

        if not stu_no:
            errors.append({'row': row_number, 'field': 'stu_no', 'message': '学号不能为空'})
        if not realname:
            errors.append({'row': row_number, 'field': 'realname', 'message': '姓名不能为空'})
        if username in existing_usernames:
            errors.append({'row': row_number, 'field': 'username', 'message': '用户名已存在'})
        if username in seen_usernames:
            errors.append({'row': row_number, 'field': 'username', 'message': 'Excel内用户名重复'})
        if stu_no in existing_stu_nos:
            errors.append({'row': row_number, 'field': 'stu_no', 'message': '学号已存在'})
        if stu_no in seen_stu_nos:
            errors.append({'row': row_number, 'field': 'stu_no', 'message': 'Excel内学号重复'})

        age = _parse_age(row_data.get('age', ''), row_number, errors)
        email = row_data.get('email', '')
        if email:
            try:
                validate_email(email)
            except DjangoValidationError:
                errors.append({'row': row_number, 'field': 'email', 'message': '邮箱格式无效'})

        gender = _normalize_gender(row_data.get('gender', ''))
        if row_data.get('gender', '') and not gender:
            errors.append({'row': row_number, 'field': 'gender', 'message': '性别只支持 男/女/MALE/FEMALE'})

        class_obj = _resolve_class(
            row_data,
            row_number,
            classes_by_id,
            classes_by_grade_name,
            errors,
        )

        seen_usernames.add(username)
        seen_stu_nos.add(stu_no)
        import_rows.append(ImportRow(
            row_number=row_number,
            username=username,
            stu_no=stu_no,
            realname=realname,
            phone=row_data.get('phone', ''),
            email=email,
            address=row_data.get('address', ''),
            age=age,
            gender=gender,
            class_obj=class_obj,
            password=row_data.get('password', '') or default_password,
        ))

    total_rows = len(import_rows)
    failed_rows = len({error['row'] for error in errors if error['row'] > 1})
    if errors:
        return {
            'total_rows': total_rows,
            'imported_count': 0,
            'failed_count': failed_rows,
            'errors': _format_errors(errors, row_context),
        }

    with transaction.atomic():
        for import_row in import_rows:
            user = User.objects.create_user(
                username=import_row.username,
                password=import_row.password,
                email=import_row.email,
                phone=import_row.phone,
                real_name=import_row.realname,
                role=RoleChoices.STUDENT,
                is_active=True,
                is_approved=True,
            )
            StudentProfile.objects.create(
                user=user,
                stu_no=import_row.stu_no,
                realname=import_row.realname,
                phone=import_row.phone,
                email=import_row.email,
                address=import_row.address,
                age=import_row.age,
                gender=import_row.gender,
                class_id=import_row.class_obj,
            )

    return {
        'total_rows': total_rows,
        'imported_count': len(import_rows),
        'failed_count': 0,
        'errors': [],
    }
