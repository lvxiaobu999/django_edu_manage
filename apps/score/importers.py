from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from apps.dicts.models import SubjectDict
from apps.exam.models import ExamPlan
from apps.score.models import Score
from apps.students.models import StudentProfile


HEADER_ALIASES = {
    'stu_no': {'stu_no', 'student_no', '学号'},
    'realname': {'realname', 'name', '姓名', '学生姓名'},
    'exam_id': {'exam_id', '考试ID', '考试id'},
    'exam_name': {'exam_name', '考试名称', '考试'},
    'subject_id': {'subject_id', '科目ID', '科目id'},
    'subject_name': {'subject_name', '科目名称', '科目'},
    'score': {'score', '分数', '成绩'},
}

TEMPLATE_HEADERS = ['学号', '姓名', '考试ID', '考试名称', '科目ID', '科目名称', '分数']


@dataclass
class ScoreImportRow:
    row_number: int
    student: StudentProfile
    exam: ExamPlan
    subject: SubjectDict
    score: Decimal


def _normalize_header(value):
    return str(value or '').strip()


def _normalize_cell(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _build_header_map(header_row):
    alias_to_field = {
        alias.lower(): field
        for field, aliases in HEADER_ALIASES.items()
        for alias in aliases
    }
    header_map = {}
    for index, value in enumerate(header_row):
        header = _normalize_header(value)
        if not header:
            continue
        field = alias_to_field.get(header.lower())
        if field:
            header_map[field] = index
    return header_map


def _row_to_dict(row, header_map):
    return {
        field: _normalize_cell(row[index] if index < len(row) else '')
        for field, index in header_map.items()
    }


def _format_errors(errors, row_context=None):
    row_context = row_context or {}
    grouped_errors = {}

    for error in errors:
        row_number = error['row']
        context = row_context.get(row_number, {})
        row_error = grouped_errors.setdefault(row_number, {
            'row': row_number,
            'stu_no': context.get('stu_no', ''),
            'realname': context.get('realname', ''),
            'messages': [],
        })
        message = error['message']
        if message not in row_error['messages']:
            row_error['messages'].append(message)

    return list(grouped_errors.values())


def _parse_decimal_score(value, row_number, errors):
    if not value:
        errors.append({'row': row_number, 'field': 'score', 'message': '分数不能为空'})
        return None

    try:
        score = Decimal(value)
    except (InvalidOperation, ValueError):
        errors.append({'row': row_number, 'field': 'score', 'message': '分数必须是数字'})
        return None

    if score < Decimal('0') or score > Decimal('999.9'):
        errors.append({'row': row_number, 'field': 'score', 'message': '分数必须在 0 到 999.9 之间'})
        return None

    rounded_score = score.quantize(Decimal('0.1'))
    if score != rounded_score:
        errors.append({'row': row_number, 'field': 'score', 'message': '分数最多保留 1 位小数'})
        return None

    return rounded_score


def _resolve_by_id_or_name(row_data, row_number, id_field, name_field, by_id, by_name, label, errors):
    raw_id = row_data.get(id_field, '')
    raw_name = row_data.get(name_field, '')

    if raw_id:
        try:
            object_id = int(raw_id)
        except ValueError:
            errors.append({'row': row_number, 'field': id_field, 'message': f'{label}ID必须是整数'})
            return None

        obj = by_id.get(object_id)
        if obj is None:
            errors.append({'row': row_number, 'field': id_field, 'message': f'{label}ID不存在'})
        return obj

    if raw_name:
        matches = by_name.get(raw_name, [])
        if len(matches) == 1:
            return matches[0]
        if len(matches) > 1:
            errors.append({'row': row_number, 'field': name_field, 'message': f'{label}名称不唯一，请填写{label}ID'})
            return None
        errors.append({'row': row_number, 'field': name_field, 'message': f'{label}名称不存在'})
        return None

    errors.append({'row': row_number, 'field': id_field, 'message': f'必须填写{label}ID或{label}名称'})
    return None


def _build_name_map(objects):
    name_map = {}
    for obj in objects:
        name_map.setdefault(obj.name, []).append(obj)
    return name_map


def build_score_import_template():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '成绩导入模板'
    worksheet.append(TEMPLATE_HEADERS)
    worksheet.append(['S20260001', '张三', 1, '示例考试名称', 1, '语文', 95.5])
    worksheet.append(['S20260002', '李四', 1, '示例考试名称', 2, '数学', 88])

    header_fill = PatternFill(fill_type='solid', fgColor='D9EAF7')
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for index, width in enumerate([14, 12, 10, 30, 10, 16, 10], start=1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = width
    worksheet.freeze_panes = 'A2'

    help_sheet = workbook.create_sheet('填写说明')
    help_rows = [
        ['项目', '说明'],
        ['必填列', '学号、分数；考试ID/考试名称二选一；科目ID/科目名称二选一'],
        ['推荐填写', '优先填写考试ID和科目ID，避免名称重复造成歧义'],
        ['分数范围', '0 到 999.9，最多保留 1 位小数'],
        ['重复成绩', '同一学生 + 同一考试 + 同一科目已存在时，默认报错；overwrite=true 时更新原成绩'],
        ['导入策略', '先校验整张表；任意行有错误则不写入任何数据'],
    ]
    for row in help_rows:
        help_sheet.append(row)
    help_sheet.column_dimensions['A'].width = 18
    help_sheet.column_dimensions['B'].width = 90
    for cell in help_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    exam_sheet = workbook.create_sheet('考试列表')
    exam_sheet.append(['考试ID', '考试名称', '年级', '考试日期'])
    for exam in ExamPlan.objects.order_by('-exam_date', 'id'):
        exam_sheet.append([exam.id, exam.name, exam.grade, exam.exam_date.isoformat()])
    for cell in exam_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for column, width in {'A': 10, 'B': 36, 'C': 14, 'D': 14}.items():
        exam_sheet.column_dimensions[column].width = width

    subject_sheet = workbook.create_sheet('科目列表')
    subject_sheet.append(['科目ID', '科目名称'])
    for subject in SubjectDict.objects.order_by('id'):
        subject_sheet.append([subject.id, subject.name])
    for cell in subject_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    subject_sheet.column_dimensions['A'].width = 10
    subject_sheet.column_dimensions['B'].width = 20

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output.getvalue()


def import_scores_from_excel(file_obj, overwrite=False):
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return {
            'total_rows': 0,
            'imported_count': 0,
            'updated_count': 0,
            'failed_count': 0,
            'errors': _format_errors([
                {'row': 1, 'field': 'file', 'message': 'Excel不能为空'},
            ]),
        }

    header_map = _build_header_map(rows[0])
    missing_columns = {'stu_no', 'score'} - set(header_map)
    if missing_columns:
        errors = [
            {'row': 1, 'field': column, 'message': '缺少必填列'}
            for column in sorted(missing_columns)
        ]
        return {
            'total_rows': 0,
            'imported_count': 0,
            'updated_count': 0,
            'failed_count': 0,
            'errors': _format_errors(errors),
        }

    students_by_no = {
        student.stu_no: student
        for student in StudentProfile.objects.select_related('user')
    }
    exams = list(ExamPlan.objects.all())
    subjects = list(SubjectDict.objects.all())
    exams_by_id = {exam.id: exam for exam in exams}
    subjects_by_id = {subject.id: subject for subject in subjects}
    exams_by_name = _build_name_map(exams)
    subjects_by_name = _build_name_map(subjects)
    existing_score_keys = set(Score.objects.values_list('student_id', 'exam_id', 'subject_id'))
    seen_score_keys = set()
    import_rows = []
    errors = []
    row_context = {}

    for row_number, row in enumerate(rows[1:], start=2):
        row_data = _row_to_dict(row, header_map)
        if not any(row_data.values()):
            continue

        stu_no = row_data.get('stu_no', '')
        realname = row_data.get('realname', '')
        row_context[row_number] = {
            'stu_no': stu_no,
            'realname': realname,
        }

        student = students_by_no.get(stu_no)
        if not stu_no:
            errors.append({'row': row_number, 'field': 'stu_no', 'message': '学号不能为空'})
        elif student is None:
            errors.append({'row': row_number, 'field': 'stu_no', 'message': '学号不存在'})
        elif realname and student.realname != realname:
            errors.append({'row': row_number, 'field': 'realname', 'message': '姓名与学号对应的学生姓名不一致'})
        elif not realname:
            row_context[row_number]['realname'] = student.realname

        exam = _resolve_by_id_or_name(
            row_data, row_number, 'exam_id', 'exam_name',
            exams_by_id, exams_by_name, '考试', errors,
        )
        subject = _resolve_by_id_or_name(
            row_data, row_number, 'subject_id', 'subject_name',
            subjects_by_id, subjects_by_name, '科目', errors,
        )
        score = _parse_decimal_score(row_data.get('score', ''), row_number, errors)

        if student and exam and subject:
            score_key = (student.id, exam.id, subject.id)
            if score_key in seen_score_keys:
                errors.append({'row': row_number, 'field': 'score', 'message': 'Excel内同一学生同一考试同一科目重复'})
            if score_key in existing_score_keys and not overwrite:
                errors.append({'row': row_number, 'field': 'score', 'message': '该学生该考试该科目成绩已存在'})
            seen_score_keys.add(score_key)

        if student and exam and subject and score is not None:
            import_rows.append(ScoreImportRow(
                row_number=row_number,
                student=student,
                exam=exam,
                subject=subject,
                score=score,
            ))

    total_rows = len({
        row_number
        for row_number, row in enumerate(rows[1:], start=2)
        if any(_normalize_cell(value) for value in row)
    })
    failed_rows = len({error['row'] for error in errors if error['row'] > 1})
    if errors:
        return {
            'total_rows': total_rows,
            'imported_count': 0,
            'updated_count': 0,
            'failed_count': failed_rows,
            'errors': _format_errors(errors, row_context),
        }

    imported_count = 0
    updated_count = 0
    with transaction.atomic():
        for import_row in import_rows:
            score_obj, created = Score.objects.update_or_create(
                student=import_row.student,
                exam=import_row.exam,
                subject=import_row.subject,
                defaults={'score': import_row.score},
            )
            if created:
                imported_count += 1
            else:
                updated_count += 1

    return {
        'total_rows': total_rows,
        'imported_count': imported_count,
        'updated_count': updated_count,
        'failed_count': 0,
        'errors': [],
    }
